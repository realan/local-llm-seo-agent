"""
Catalog processing pipeline for souvenir SEO enrichment.
"""

import json
import logging
import shutil
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError, field_validator

from app.llm.ollama_client import OllamaClient
from app.services.symbolism_service import SymbolismService
from app.tools.catalog_tools import (
    build_base_tags,
    classify_size,
    detect_entity_type,
)
from app.tools.excel_tools import normalize_column_name

logger = logging.getLogger(__name__)

OUTPUT_COLUMNS = [
    "entity_type",
    "entity_confidence",
    "size_tag",
    "size_reason",
    "symbolism_summary",
    "symbolism_source_note",
    "seo_keywords",
    "seo_title",
    "seo_description",
    "product_description",
    "processed_status",
    "processed_error",
]


class SeoFields(BaseModel):
    """Structured SEO fields returned by the LLM."""

    seo_keywords: str = Field(..., max_length=300)
    seo_title: str = Field(..., max_length=80)
    seo_description: str = Field(..., max_length=180)
    product_description: str

    @field_validator("seo_keywords")
    @classmethod
    def validate_keywords(cls, value: str) -> str:
        if "," not in value:
            raise ValueError("seo_keywords must be comma separated")
        return value.strip()

    @field_validator("product_description")
    @classmethod
    def validate_description_length(cls, value: str) -> str:
        text = value.strip()
        if len(text) < 350 or len(text) > 650:
            raise ValueError("product_description must be 350-650 chars")
        return text


class CatalogProcessor:
    """Process catalog workbook row by row and write SEO enrichment."""

    def __init__(
        self,
        llm_client: OllamaClient,
        input_path: str,
        output_path: str,
        sheet_name: str = "products",
        limit: Optional[int] = None,
        cache_path: Optional[str] = None,
    ):
        self.llm_client = llm_client
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)
        self.sheet_name = sheet_name
        self.limit = limit
        self.symbolism_service = SymbolismService(cache_path=cache_path)

    def process(self) -> dict[str, Any]:
        """Run full catalog enrichment pipeline."""
        if not self.input_path.exists():
            raise ValueError(f"Input file not found: {self.input_path}")

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.input_path, self.output_path)

        workbook = load_workbook(self.output_path)
        try:
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {self.sheet_name}")

            sheet = workbook[self.sheet_name]
            headers = self._ensure_output_columns(sheet)
            self._require_column(headers, "name")

            processed = 0
            stats = {"success": 0, "skipped": 0, "needs_review": 0, "error": 0}

            for excel_row in range(2, sheet.max_row + 1):
                if self.limit is not None and processed >= self.limit:
                    break
                row = self._read_row(sheet, headers, excel_row)
                fields = self._process_row(row)
                self._write_row_fields(sheet, headers, excel_row, fields)
                status = fields.get("processed_status", "error")
                stats[status] = stats.get(status, 0) + 1
                processed += 1

            workbook.save(self.output_path)
            return {
                "processed_rows": processed,
                "output_path": str(self.output_path),
                "stats": stats,
            }
        finally:
            workbook.close()

    def _ensure_output_columns(self, sheet) -> dict[str, int]:
        """Ensure output columns exist and return normalized header map."""
        headers: dict[str, int] = {}
        for idx, cell in enumerate(sheet[1], start=1):
            name = normalize_column_name(cell.value)
            if name:
                headers[name] = idx

        next_col = sheet.max_column + 1
        for output_name in OUTPUT_COLUMNS:
            normalized = normalize_column_name(output_name)
            if normalized not in headers:
                sheet.cell(row=1, column=next_col).value = output_name
                headers[normalized] = next_col
                next_col += 1
        return headers

    @staticmethod
    def _require_column(headers: dict[str, int], column_name: str) -> None:
        if column_name not in headers:
            raise ValueError(f"Required column missing: {column_name}")

    @staticmethod
    def _read_row(sheet, headers: dict[str, int], excel_row: int) -> dict[str, Any]:
        row = {}
        for name, col_idx in headers.items():
            row[name] = sheet.cell(row=excel_row, column=col_idx).value
        return row

    def _process_row(self, row: dict[str, Any]) -> dict[str, Any]:
        if not any(str(value).strip() for value in row.values() if value is not None):
            return {"processed_status": "skipped", "processed_error": ""}

        name = str(row.get("name") or "").strip()
        if not name:
            return {"processed_status": "error", "processed_error": "missing_name"}

        category = str(row.get("category") or "").strip()
        material = str(row.get("material") or "").strip()
        article = str(row.get("article") or "").strip()

        entity = detect_entity_type(name=name, category=category)
        size = classify_size(
            name=name,
            height_cm=row.get("height_cm"),
            weight_g=row.get("weight_g"),
        )

        base_fields = {
            "entity_type": entity["entity_type"],
            "entity_confidence": entity["confidence"],
            "size_tag": size["size_tag"],
            "size_reason": size["size_reason"],
        }

        if entity["entity_type"] == "unknown":
            return {
                **base_fields,
                "processed_status": "needs_review",
                "processed_error": "entity_not_detected",
            }

        try:
            symbolism = self.symbolism_service.get_symbolism(
                entity_type=entity["entity_type"],
                name=name,
                category=category,
                article=article,
            )
        except ValueError as e:
            return {
                **base_fields,
                "processed_status": "error",
                "processed_error": str(e),
            }

        try:
            seo = self._generate_seo_fields(
                name=name,
                category=category,
                material=material,
                article=article,
                entity_type=entity["entity_type"],
                entity_confidence=entity["confidence"],
                size_tag=size["size_tag"],
                size_reason=size["size_reason"],
                symbolism_summary=symbolism["summary"],
                symbolism_keywords=symbolism["keywords"],
                base_tags=build_base_tags(
                    entity_type=entity["entity_type"],
                    size_tag=size["size_tag"],
                    material=material,
                    category=category,
                    article=article,
                ),
            )
        except ValueError as e:
            return {
                **base_fields,
                "symbolism_summary": symbolism["summary"],
                "symbolism_source_note": symbolism["source_note"],
                "processed_status": "error",
                "processed_error": str(e),
            }

        status = "success" if entity["confidence"] == "high" else "needs_review"
        return {
            **base_fields,
            "symbolism_summary": symbolism["summary"],
            "symbolism_source_note": symbolism["source_note"],
            "seo_keywords": seo.seo_keywords,
            "seo_title": seo.seo_title,
            "seo_description": seo.seo_description,
            "product_description": seo.product_description,
            "processed_status": status,
            "processed_error": "",
        }

    def _generate_seo_fields(self, **payload: Any) -> SeoFields:
        """Generate SEO fields with structured LLM output and one repair attempt."""
        prompt = self._build_seo_prompt(payload)
        response_text = self.llm_client.generate(prompt=prompt, temperature=0.4)
        response_dict = self.llm_client.parse_json_response(response_text)
        if response_dict is None:
            repair_prompt = (
                "Your previous response was invalid. Return only JSON with keys "
                "seo_keywords, seo_title, seo_description, product_description."
            )
            response_text = self.llm_client.generate(
                prompt=f"{prompt}\n\n{repair_prompt}",
                temperature=0.2,
            )
            response_dict = self.llm_client.parse_json_response(response_text)

        if response_dict is None:
            raise ValueError("invalid_model_output")

        try:
            return SeoFields(**response_dict)
        except ValidationError as e:
            logger.error("Invalid SEO fields: %s", e)
            raise ValueError("invalid_model_output") from e

    @staticmethod
    def _build_seo_prompt(payload: dict[str, Any]) -> str:
        """Build strict JSON prompt for SEO generation."""
        context = json.dumps(payload, ensure_ascii=False, indent=2)
        return (
            "You generate SEO fields for a Russian souvenir product card.\n"
            "Return only valid JSON with keys: seo_keywords, seo_title, "
            "seo_description, product_description.\n"
            "Rules:\n"
            "- seo_keywords: comma-separated string\n"
            "- seo_title: max 80 chars\n"
            "- seo_description: max 180 chars, 1-2 sentences, neutral tone\n"
            "- product_description: 350-650 chars, suitable for ecommerce\n"
            "- Mention entity type, size, material, and short neutral symbolism\n"
            "- Do not invent facts and do not use SKU\n"
            "- Neutral wording only\n\n"
            f"Product context:\n{context}"
        )

    @staticmethod
    def _write_row_fields(sheet, headers: dict[str, int], excel_row: int, fields: dict[str, Any]) -> None:
        """Write processed fields to worksheet."""
        for field_name, value in fields.items():
            normalized = normalize_column_name(field_name)
            col_idx = headers[normalized]
            sheet.cell(row=excel_row, column=col_idx).value = value
