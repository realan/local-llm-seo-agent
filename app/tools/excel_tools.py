"""
Excel tools for catalog processing.
"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.tools.base import BaseTool

logger = logging.getLogger(__name__)


def normalize_column_name(value: Any) -> str:
    """Normalize Excel header names to stable snake_case identifiers."""
    text = str(value or "").strip().lower().replace("ё", "е")
    text = text.replace("\n", " ").replace("-", "_").replace(" ", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def get_sheet_and_headers(file_path: str, sheet_name: str):
    """Open workbook and return worksheet with normalized header mapping."""
    workbook = load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    headers: dict[str, int] = {}
    for idx, cell in enumerate(sheet[1], start=1):
        normalized = normalize_column_name(cell.value)
        if normalized:
            headers[normalized] = idx
    return workbook, sheet, headers


class GetExcelInfoTool(BaseTool):
    """Return sheet information for catalog workbook."""

    def __init__(self):
        super().__init__("get_excel_info", "return sheet name, columns, and row count")

    def validate_inputs(self, **kwargs) -> bool:
        file_path = kwargs.get("file_path", "")
        sheet_name = kwargs.get("sheet_name", "")
        if not isinstance(file_path, str) or not file_path.strip():
            raise ValueError("file_path is required")
        if not isinstance(sheet_name, str) or not sheet_name.strip():
            raise ValueError("sheet_name is required")
        return True

    def run(self, file_path: str, sheet_name: str) -> dict[str, Any]:
        workbook, sheet, headers = get_sheet_and_headers(file_path, sheet_name)
        try:
            return {
                "sheet_name": sheet.title,
                "columns": list(headers.keys()),
                "row_count": max(sheet.max_row - 1, 0),
            }
        finally:
            workbook.close()

    def get_input_schema(self) -> dict[str, Any]:
        return {
            "file_path": {
                "type": "string",
                "description": "Path to .xlsx file, e.g. samples/products.xlsx",
            },
            "sheet_name": {
                "type": "string",
                "description": "Worksheet name, e.g. products",
            },
        }


class ReadExcelRowTool(BaseTool):
    """Read one normalized row from workbook."""

    def __init__(self):
        super().__init__("read_excel_row", "read one row from Excel")

    def validate_inputs(self, **kwargs) -> bool:
        file_path = kwargs.get("file_path", "")
        sheet_name = kwargs.get("sheet_name", "")
        row_index = kwargs.get("row_index", None)
        if not isinstance(file_path, str) or not file_path.strip():
            raise ValueError("file_path is required")
        if not isinstance(sheet_name, str) or not sheet_name.strip():
            raise ValueError("sheet_name is required")
        if not isinstance(row_index, int) or row_index < 0:
            raise ValueError("row_index must be a non-negative integer")
        return True

    def run(self, file_path: str, sheet_name: str, row_index: int) -> dict[str, Any]:
        workbook, sheet, headers = get_sheet_and_headers(file_path, sheet_name)
        try:
            excel_row = row_index + 2
            if excel_row > sheet.max_row:
                raise ValueError(f"Row out of range: {row_index}")

            row = {}
            for name, col_idx in headers.items():
                row[name] = sheet.cell(row=excel_row, column=col_idx).value
            return {"row": row}
        finally:
            workbook.close()

    def get_input_schema(self) -> dict[str, Any]:
        return {
            "file_path": {
                "type": "string",
                "description": "Path to .xlsx file, e.g. samples/products.xlsx",
            },
            "sheet_name": {
                "type": "string",
                "description": "Worksheet name, e.g. products",
            },
            "row_index": {
                "type": "integer",
                "description": "Zero-based data row index, e.g. 0 for the first row after headers",
            },
        }


class WriteExcelRowFieldsTool(BaseTool):
    """Write output fields into a workbook row."""

    def __init__(self):
        super().__init__(
            "write_excel_row_fields",
            "write generated fields into one row of output Excel file",
        )

    def validate_inputs(self, **kwargs) -> bool:
        input_file_path = kwargs.get("input_file_path", "")
        output_file_path = kwargs.get("output_file_path", "")
        sheet_name = kwargs.get("sheet_name", "")
        row_index = kwargs.get("row_index", None)
        fields = kwargs.get("fields", None)
        if not isinstance(input_file_path, str) or not input_file_path.strip():
            raise ValueError("input_file_path is required")
        if not isinstance(output_file_path, str) or not output_file_path.strip():
            raise ValueError("output_file_path is required")
        if not isinstance(sheet_name, str) or not sheet_name.strip():
            raise ValueError("sheet_name is required")
        if not isinstance(row_index, int) or row_index < 0:
            raise ValueError("row_index must be a non-negative integer")
        if not isinstance(fields, dict):
            raise ValueError("fields must be an object")
        return True

    def run(
        self,
        input_file_path: str,
        output_file_path: str,
        sheet_name: str,
        row_index: int,
        fields: dict[str, Any],
    ) -> dict[str, Any]:
        output = Path(output_file_path)
        if not output.exists():
            raise ValueError(
                f"Output workbook does not exist yet: {output_file_path}. "
                f"Create the output copy before writing rows."
            )

        workbook, sheet, headers = get_sheet_and_headers(str(output), sheet_name)
        try:
            excel_row = row_index + 2
            next_col = sheet.max_column + 1
            for field_name, value in fields.items():
                normalized = normalize_column_name(field_name)
                col_idx = headers.get(normalized)
                if col_idx is None:
                    col_idx = next_col
                    next_col += 1
                    headers[normalized] = col_idx
                    sheet.cell(row=1, column=col_idx).value = field_name
                sheet.cell(row=excel_row, column=col_idx).value = value
            workbook.save(output_file_path)
            return {"success": True}
        finally:
            workbook.close()

    def get_input_schema(self) -> dict[str, Any]:
        return {
            "input_file_path": {"type": "string", "description": "Original workbook path"},
            "output_file_path": {"type": "string", "description": "Writable output workbook path"},
            "sheet_name": {"type": "string", "description": "Worksheet name"},
            "row_index": {"type": "integer", "description": "Zero-based data row index"},
            "fields": {"type": "object", "description": "Fields to write into the row"},
        }
