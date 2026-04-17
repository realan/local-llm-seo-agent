import json

from openpyxl import Workbook, load_workbook

from app.services.catalog_processor import CatalogProcessor
from app.tools.catalog_tools import classify_size, detect_entity_type


class FakeLlmClient:
    def generate(self, prompt: str, temperature: float = 0.4, top_p: float = 0.9, top_k: int = 40) -> str:
        return json.dumps(
            {
                "seo_keywords": "слон, фигурка слона, сувенир слон, фарфор",
                "seo_title": "Фигурка слона малая из фарфора",
                "seo_description": "Нейтральное описание сувенирной фигурки слона малого размера из фарфора.",
                "product_description": (
                    "Фигурка слона малого размера из фарфора подходит для декоративного оформления "
                    "интерьера и подарочного ассортимента. Такой образ традиционно связывают с "
                    "мудростью, устойчивостью и спокойным характером, поэтому изделие легко "
                    "вписывается в коллекции сувениров и тематических композиций. Компактный формат "
                    "удобен для полки, витрины или рабочего стола, а нейтральная подача делает товар "
                    "подходящим для разных сценариев покупки."
                ),
            },
            ensure_ascii=False,
        )

    def parse_json_response(self, text: str):
        return json.loads(text)


class FakeResponse:
    status_code = 200
    text = '{"ok": true}'

    def raise_for_status(self):
        return None


def test_detect_entity_type_alias_high_confidence():
    result = detect_entity_type(name="Фигурка слоник", category="Сувенир")
    assert result["entity_type"] == "elephant"
    assert result["confidence"] in {"high", "medium"}


def test_classify_size_prefers_title_then_numeric_conflict():
    result = classify_size(name="Слон малый", height_cm=20, weight_g=1200)
    assert result["size_tag"] == "large"
    assert "title_hint_small" in result["size_reason"]
    assert "weight_large" in result["size_reason"]


def test_catalog_processor_writes_output_fields(tmp_path, monkeypatch):
    input_path = tmp_path / "products.xlsx"
    output_path = tmp_path / "products_result.xlsx"
    cache_path = tmp_path / "symbolism_cache.json"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products"
    sheet.append(["name", "category", "height_cm", "weight_g", "material", "article"])
    sheet.append(["Слоник малый", "сувенир", 6, 80, "фарфор", "ART-1"])
    sheet.append(["Неясная фигурка", "декор", None, None, "смола", "ART-2"])
    sheet.append(["   ", None, None, None, None, None])
    workbook.save(input_path)

    monkeypatch.setattr("app.tools.http_fetch.requests.get", lambda *args, **kwargs: FakeResponse())

    processor = CatalogProcessor(
        llm_client=FakeLlmClient(),
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name="products",
        cache_path=str(cache_path),
    )

    result = processor.process()

    assert result["processed_rows"] == 3
    assert result["stats"]["success"] == 1
    assert result["stats"]["needs_review"] == 1
    assert result["stats"]["skipped"] == 1

    out_book = load_workbook(output_path)
    out_sheet = out_book["products"]
    headers = [cell.value for cell in out_sheet[1]]
    row1 = {headers[idx]: out_sheet.cell(row=2, column=idx + 1).value for idx in range(len(headers))}
    row2 = {headers[idx]: out_sheet.cell(row=3, column=idx + 1).value for idx in range(len(headers))}
    row3 = {headers[idx]: out_sheet.cell(row=4, column=idx + 1).value for idx in range(len(headers))}
    out_book.close()

    assert row1["entity_type"] == "elephant"
    assert row1["size_tag"] == "small"
    assert row1["processed_status"] == "success"
    assert row1["seo_title"] == "Фигурка слона малая из фарфора"
    assert row1["symbolism_source_note"].startswith("http_lookup:")

    assert row2["processed_status"] == "needs_review"
    assert row2["processed_error"] == "entity_not_detected"

    assert row3["processed_status"] == "skipped"
